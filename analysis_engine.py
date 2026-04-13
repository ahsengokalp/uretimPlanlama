import unicodedata
import warnings
from decimal import Decimal, ROUND_HALF_UP

import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook


PLANLAMA_RULES = [
    ("fatura bekleyen hazir ton", 2, "Girisler neden aksiyor? Ne yapacagiz?"),
    ("kalite kontrol bekleyen ton", 1, "Kalite kontrolde bekleyen isler icin karar hizlanmali."),
    ("rework bekleyen ton", 1, "Rework akisini hizlandiracak net bir aksiyon belirlenmeli."),
    ("taslama stogu ton", 20, "Taslama stogu kritik seviyede. Yuku dagitacak plan gerekli."),
    ("kumlama stogu ton", 20, "Kumlama stogu kritik seviyede. Birikme noktasi incelenmeli."),
    ("dokulecek is miktari ton", 50, "Dokum yuku yuksek. Hat ve mesai plani gozden gecirilmeli."),
    ("mevcut siparis ton", 20, "Mevcut siparis yuku kritik seviyeyi asti."),
    ("2 hafta sonrasi icin dokum eksigi ton", 30, "Iki hafta sonrasi icin dokum eksigi kritik seviyede."),
    ("1 hafta sonrasi icin dokum eksigi ton", 10, "Bir hafta sonrasi icin dokum eksigi kritik seviyede."),
    ("mevcut hafta icin dokum eksigi ton", 5, "Mevcut hafta icin dokum eksigi kritik seviyede."),
    ("bakiye icin dokum eksigi ton", 2, "Bakiye dokum eksigi icin kok neden kontrol edilmeli."),
]

CATEGORY_ORDER = {
    "isg": 0,
    "kalite": 1,
    "uretim": 2,
    "planlama": 3,
}

STATUS_ORDER = {
    "danger": 0,
    "warning": 1,
    "success": 2,
    "info": 3,
}


def normalize_columns(df):
    df.columns = [str(col).strip() if not isinstance(col, pd.Timestamp) else col for col in df.columns]
    return df


def normalize_label(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = text.replace("\u0131", "i").replace("\u0130", "I")
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.strip().lower().split())


def get_category_kind(value):
    key = normalize_label(value)
    if "planlama" in key:
        return "planlama"
    if "kalite" in key:
        return "kalite"
    if "uretim" in key:
        return "uretim"
    if "isg" in key:
        return "isg"
    return key


def infer_unit(parametre):
    original = str(parametre)
    normalized = normalize_label(parametre)

    if "%" in original or " oran" in normalized or "yuzde" in normalized:
        return "%"
    if " ton" in normalized:
        return "T"
    if " kg" in normalized:
        return "kg"
    if "adet" in normalized or "sayi" in normalized or "kalip" in normalized:
        return "adet"
    return None


def normalize_percent_value(value, parametre):
    if pd.isna(value):
        return value
    if infer_unit(parametre) != "%":
        return float(value)
    return float(value)


def to_display_number(value, decimals=1):
    if pd.isna(value):
        return "-"

    value = float(value)
    rounded = round(value, decimals)
    if abs(rounded - round(rounded)) < 1e-9:
        return str(int(round(rounded)))
    return f"{rounded:.{decimals}f}".rstrip("0").rstrip(".")


def decimal_places_from_format(number_format):
    if not number_format:
        return 0

    format_text = str(number_format).split(";")[0]
    if "." not in format_text:
        return 0

    decimal_part = format_text.split(".", 1)[1]
    return sum(1 for char in decimal_part if char in {"0", "#"})


def quantize_decimal(value, decimals):
    quantizer = Decimal("1") if decimals <= 0 else Decimal(f"1.{'0' * decimals}")
    return Decimal(str(value)).quantize(quantizer, rounding=ROUND_HALF_UP)


def format_excel_cell_value(value, number_format, parametre=None):
    if value is None or pd.isna(value):
        return None

    format_text = str(number_format or "").split(";")[0]
    decimals = decimal_places_from_format(format_text)

    if "%" in format_text:
        percent_value = quantize_decimal(float(value) * 100, decimals)
        return f"{percent_value:.{decimals}f}%" if decimals > 0 else f"{int(percent_value)}%"

    unit = infer_unit(parametre) if parametre else None

    if any(token in format_text for token in ("0", "#")):
        numeric_value = quantize_decimal(float(value), decimals)
        display_value = f"{numeric_value:.{decimals}f}" if decimals > 0 else str(int(numeric_value))
        if unit == "%":
            return f"{display_value}%"
        if unit == "T":
            return f"{display_value}T"
        if unit == "kg":
            return f"{display_value} kg"
        if unit == "adet":
            return f"{display_value} adet"
        return display_value

    numeric_value = pd.to_numeric(value, errors="coerce")
    if not pd.isna(numeric_value) and unit == "%":
        return f"{to_display_number(float(numeric_value), decimals=1)}%"

    return str(value)


def build_excel_display_map(filepath, sheet_name="Veriler"):
    display_map = {}

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
        )
        workbook = load_workbook(filepath, data_only=True)

    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        date_columns = []

        for column_index, header_value in enumerate(headers[2:], start=3):
            parsed = pd.to_datetime(header_value, errors="coerce")
            if not pd.isna(parsed):
                date_columns.append((column_index, pd.Timestamp(parsed).normalize()))

        for row in worksheet.iter_rows(min_row=2):
            kategori = row[0].value
            parametre = row[1].value
            if kategori is None or parametre is None:
                continue

            kategori_key = normalize_label(kategori)
            parametre_key = normalize_label(parametre)

            for column_index, tarih in date_columns:
                cell = worksheet.cell(row=row[0].row, column=column_index)
                if cell.value is None:
                    continue

                display_value = format_excel_cell_value(cell.value, cell.number_format, parametre)
                if display_value is None:
                    continue

                display_map[(kategori_key, parametre_key, tarih)] = display_value
    finally:
        workbook.close()

    return display_map


def format_metric_value(value, parametre):
    if pd.isna(value):
        return "-"

    unit = infer_unit(parametre)
    if unit == "%":
        percent_value = normalize_percent_value(value, parametre)
        return f"{percent_value:.1f}%"

    display_value = to_display_number(value, decimals=1)
    if unit == "T":
        return f"{display_value}T"
    if unit == "kg":
        return f"{display_value} kg"
    if unit == "adet":
        return f"{display_value} adet"
    return display_value


def format_comparison(value, target, parametre, value_display=None, target_display=None):
    if pd.isna(value) or pd.isna(target):
        return None
    operator = ">" if float(value) > float(target) else "<" if float(value) < float(target) else "="
    left_text = value_display or format_metric_value(value, parametre)
    right_text = target_display or format_metric_value(target, parametre)
    return f"{left_text} {operator} {right_text}"


def prepare_dataframe(df):
    df = normalize_columns(df)

    if "Kategori" not in df.columns or "Parametre" not in df.columns:
        raise ValueError("Veriler sayfasinda 'Kategori' ve 'Parametre' sutunlari bulunamadi.")

    candidate_columns = [col for col in df.columns if col not in ["Kategori", "Parametre"]]

    date_columns = []
    for col in candidate_columns:
        parsed = pd.to_datetime(col, errors="coerce")
        if not pd.isna(parsed):
            date_columns.append(col)

    if not date_columns:
        raise ValueError(f"Tarih sutunlari bulunamadi. Bulunan sutunlar: {list(df.columns)}")

    melted = df.melt(
        id_vars=["Kategori", "Parametre"],
        value_vars=date_columns,
        var_name="Tarih",
        value_name="Deger",
    )

    melted["Tarih"] = pd.to_datetime(melted["Tarih"], errors="coerce")
    melted["Deger"] = pd.to_numeric(melted["Deger"], errors="coerce")
    melted["Kategori"] = melted["Kategori"].astype(str).str.strip()
    melted["Parametre"] = melted["Parametre"].astype(str).str.strip()

    melted = melted.dropna(subset=["Tarih", "Deger"]).copy()
    return melted


def enrich_with_excel_display(df, display_map):
    enriched = df.copy()
    enriched["DegerGosterimExcel"] = [
        display_map.get(
            (
                normalize_label(kategori),
                normalize_label(parametre),
                pd.Timestamp(tarih).normalize(),
            )
        )
        for kategori, parametre, tarih in zip(enriched["Kategori"], enriched["Parametre"], enriched["Tarih"])
    ]
    return enriched


def clean_base_name(parametre):
    value = str(parametre).strip()
    lowered = value.lower()
    replacements = [
        (" hedefi %", " %"),
        (" hedef %", " %"),
        (" hedefi", ""),
        (" hedef", ""),
    ]

    for suffix, add_back in replacements:
        if lowered.endswith(suffix):
            base = value[: len(value) - len(suffix)].strip()
            return f"{base}{add_back}".strip()
    return value


def attach_targets(df):
    temp = df.copy()
    temp["KategoriKey"] = temp["Kategori"].apply(get_category_kind)
    temp["BaseParametre"] = temp["Parametre"].apply(clean_base_name)
    temp["BaseKey"] = temp["BaseParametre"].apply(normalize_label)

    is_target = temp["Parametre"].apply(lambda value: "hedef" in normalize_label(value))

    target_df = temp[is_target].copy()
    actual_df = temp[~is_target].copy()

    rename_map = {"Deger": "Hedef"}
    if "DegerGosterimExcel" in target_df.columns:
        rename_map["DegerGosterimExcel"] = "HedefGosterimExcel"
    target_df = target_df.rename(columns=rename_map)

    target_columns = ["KategoriKey", "BaseKey", "Tarih", "Hedef"]
    if "HedefGosterimExcel" in target_df.columns:
        target_columns.append("HedefGosterimExcel")
    target_df = target_df[target_columns]

    actual_df = actual_df.merge(
        target_df,
        how="left",
        on=["KategoriKey", "BaseKey", "Tarih"],
    )

    return actual_df.sort_values(["Kategori", "Parametre", "Tarih"]).reset_index(drop=True)


def get_meeting_dates(df):
    operational_df = df[df["Kategori"].apply(lambda value: get_category_kind(value) in {"isg", "kalite", "uretim"})].copy()
    if operational_df.empty:
        raise ValueError("ISG / Kalite / Uretim icin tarih bulunamadi.")
    operational_day = operational_df["Tarih"].max().normalize()

    planlama_df = df[df["Kategori"].apply(lambda value: get_category_kind(value) == "planlama")].copy()
    planlama_day = planlama_df["Tarih"].max().normalize() if not planlama_df.empty else operational_day

    return planlama_day, operational_day


def build_action_payload(
    row,
    day,
    yorum,
    comparison_value=None,
    status="info",
    status_label=None,
    relation_label=None,
):
    hedef = row.get("Hedef")
    deger = row["Deger"]
    deger_gosterim = row.get("DegerGosterimExcel") or format_metric_value(deger, row["Parametre"])
    hedef_gosterim = None
    if not pd.isna(hedef):
        hedef_gosterim = row.get("HedefGosterimExcel") or format_metric_value(hedef, row["Parametre"])

    if comparison_value is None:
        comparison_text = format_comparison(
            deger,
            hedef,
            row["Parametre"],
            value_display=deger_gosterim,
            target_display=hedef_gosterim,
        )
    else:
        comparison_text = format_comparison(
            deger,
            comparison_value,
            row["Parametre"],
            value_display=deger_gosterim,
            target_display=format_metric_value(comparison_value, row["Parametre"]),
        )

    return {
        "inceleme_gunu": str(day.date()),
        "kategori": row["Kategori"],
        "kategori_tipi": get_category_kind(row["Kategori"]),
        "parametre": row["Parametre"],
        "status": status,
        "status_label": status_label or "Inceleme",
        "relation_label": relation_label,
        "deger": deger,
        "hedef": hedef if not pd.isna(hedef) else None,
        "deger_gosterim": deger_gosterim,
        "hedef_gosterim": hedef_gosterim,
        "karsilastirma": comparison_text,
        "yorum": yorum,
    }


def generate_planlama_action(row):
    normalized_param = normalize_label(row["Parametre"])
    value = float(row["Deger"])

    for key, threshold, message in PLANLAMA_RULES:
        if key in normalized_param and value > threshold:
            payload = build_action_payload(
                row,
                row["Tarih"].normalize(),
                message,
                comparison_value=threshold,
                status="danger",
                status_label="Kritik",
                relation_label="Limitin ustunde",
            )
            payload["hedef"] = threshold
            payload["hedef_gosterim"] = format_metric_value(threshold, row["Parametre"])
            payload["karsilastirma"] = f"{format_metric_value(value, row['Parametre'])} > {format_metric_value(threshold, row['Parametre'])}"
            return payload

    return None


def generate_isg_action(row):
    if "kaza" not in normalize_label(row["Parametre"]):
        return None
    if float(row["Deger"]) > 0:
        return build_action_payload(
            row,
            row["Tarih"].normalize(),
            "Kaza nedeni netlestirilmeli ve aksiyon plani acilmali.",
            comparison_value=0,
            status="danger",
            status_label="Kritik",
            relation_label="Kaza var",
        )
    return build_action_payload(
        row,
        row["Tarih"].normalize(),
        "Boyle devam!",
        comparison_value=0,
        status="success",
        status_label="Iyi",
        relation_label="Sifir kaza",
    )


def generate_kalite_uretim_action(row):
    hedef = row.get("Hedef")
    if pd.isna(hedef):
        return None

    kategori = get_category_kind(row["Kategori"])
    deger = float(row["Deger"])
    hedef = float(hedef)

    if kategori == "kalite" and deger > hedef:
        return build_action_payload(
            row,
            row["Tarih"].normalize(),
            "Kalite hedefi asildi. Hangi hat veya urunde sapma oldugu incelenmeli.",
            status="danger",
            status_label="Kritik",
            relation_label="Hedefin ustunde",
        )

    if kategori == "kalite":
        relation_label = "Hedefe esit" if deger == hedef else "Hedefin altinda"
        return build_action_payload(
            row,
            row["Tarih"].normalize(),
            "Tebrikler. Hedefleneni gerceklestirdin.",
            status="success",
            status_label="Iyi",
            relation_label=relation_label,
        )

    if kategori == "uretim" and deger < hedef:
        return build_action_payload(
            row,
            row["Tarih"].normalize(),
            "Uretim hedefin altinda kaldi. Kapasite ve durus nedenleri kontrol edilmeli.",
            status="danger",
            status_label="Kritik",
            relation_label="Hedefin altinda",
        )

    if kategori == "uretim":
        relation_label = "Hedefe esit" if deger == hedef else "Hedefin ustunde"
        return build_action_payload(
            row,
            row["Tarih"].normalize(),
            "Tebrikler. Bunu nasil surdurebiliriz?",
            status="success",
            status_label="Iyi",
            relation_label=relation_label,
        )

    return None


def build_daily_review(df):
    planlama_day, operational_day = get_meeting_dates(df)
    review_rows = []

    operational_df = df[
        (df["Kategori"].apply(lambda value: get_category_kind(value) in {"isg", "kalite", "uretim"}))
        & (df["Tarih"].dt.normalize() == operational_day)
    ].copy()

    for _, row in operational_df.iterrows():
        kategori = get_category_kind(row["Kategori"])
        payload = None

        if kategori == "isg":
            payload = generate_isg_action(row)
        elif kategori in {"kalite", "uretim"}:
            payload = generate_kalite_uretim_action(row)

        if payload:
            payload["inceleme_gunu"] = str(operational_day.date())
            review_rows.append(payload)

    planlama_df = df[
        (df["Kategori"].apply(lambda value: get_category_kind(value) == "planlama"))
        & (df["Tarih"].dt.normalize() == planlama_day)
    ].copy()

    for _, row in planlama_df.iterrows():
        payload = generate_planlama_action(row)
        if payload:
            payload["inceleme_gunu"] = str(planlama_day.date())
            review_rows.append(payload)

    review_rows.sort(
        key=lambda item: (
            CATEGORY_ORDER.get(item["kategori_tipi"], 99),
            STATUS_ORDER.get(item["status"], 99),
            item["parametre"],
        )
    )

    return review_rows, planlama_day, operational_day


def build_parameter_summaries(df):
    summaries = []

    grouped = df.groupby(["Kategori", "Parametre"], sort=True)
    for (kategori, parametre), group in grouped:
        group = group.sort_values("Tarih")
        latest_row = group.iloc[-1]
        latest_target = group["Hedef"].dropna()
        target_value = latest_target.iloc[-1] if not latest_target.empty else None
        latest_value_display = latest_row.get("DegerGosterimExcel") or format_metric_value(latest_row["Deger"], parametre)
        latest_target_display = latest_row.get("HedefGosterimExcel") or (
            format_metric_value(target_value, parametre) if target_value is not None else None
        )

        summaries.append(
            {
                "kategori": kategori,
                "kategori_tipi": get_category_kind(kategori),
                "parametre": parametre,
                "son_tarih": str(latest_row["Tarih"].date()),
                "guncel_deger": latest_row["Deger"],
                "guncel_deger_gosterim": latest_value_display,
                "ortalama": group["Deger"].mean(),
                "ortalama_gosterim": format_metric_value(group["Deger"].mean(), parametre),
                "maksimum": group["Deger"].max(),
                "maksimum_gosterim": format_metric_value(group["Deger"].max(), parametre),
                "minimum": group["Deger"].min(),
                "minimum_gosterim": format_metric_value(group["Deger"].min(), parametre),
                "hedef": target_value,
                "hedef_gosterim": latest_target_display,
                "karsilastirma": format_comparison(
                    latest_row["Deger"],
                    target_value,
                    parametre,
                    value_display=latest_value_display,
                    target_display=latest_target_display,
                )
                if target_value is not None
                else None,
            }
        )

    summaries.sort(key=lambda item: (item["kategori_tipi"], item["parametre"]))
    return summaries


def build_highlight_actions(actions):
    ordered = sorted(
        actions,
        key=lambda item: (
            STATUS_ORDER.get(item["status"], 99),
            CATEGORY_ORDER.get(item["kategori_tipi"], 99),
            item["parametre"],
        ),
    )
    return ordered[:6]


def scale_series_for_chart(series, parametre):
    unit = infer_unit(parametre)
    return series, unit


def create_charts(df):
    charts = []

    grouped = df.groupby(["Kategori", "Parametre"], sort=True)
    for (kategori, parametre), group in grouped:
        group = group.sort_values("Tarih")
        values, unit = scale_series_for_chart(group["Deger"], parametre)

        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=group["Tarih"],
                y=values,
                mode="lines+markers",
                name="Gerceklesen",
                line={"color": "#0f6af2", "width": 3},
                marker={"size": 7},
            )
        )

        if group["Hedef"].notna().any():
            targets, _ = scale_series_for_chart(group["Hedef"], parametre)
            fig.add_trace(
                go.Scatter(
                    x=group["Tarih"],
                    y=targets,
                    mode="lines",
                    name="Hedef",
                    line={"color": "#ea4f5b", "width": 2, "dash": "dash"},
                )
            )

        axis_suffix = unit if unit in {"%", "T"} else ""
        fig.update_layout(
            margin={"l": 24, "r": 20, "t": 16, "b": 24},
            height=320,
            paper_bgcolor="#ffffff",
            plot_bgcolor="#ffffff",
            legend={"orientation": "h", "yanchor": "bottom", "y": 1.02, "x": 0},
            xaxis={"showgrid": False},
            yaxis={"gridcolor": "#e6ecf5", "ticksuffix": axis_suffix},
        )

        charts.append(
            {
                "title": parametre,
                "subtitle": kategori,
                "category_key": get_category_kind(kategori),
                "html": fig.to_html(
                    full_html=False,
                    include_plotlyjs="cdn",
                    config={
                        "responsive": True,
                        "displaylogo": False,
                        "modeBarButtonsToRemove": ["lasso2d", "select2d", "autoScale2d"],
                    },
                ),
            }
        )

    return charts


def analyze_raw_dataframe(raw_df, display_map=None):
    df = prepare_dataframe(raw_df)
    if display_map:
        df = enrich_with_excel_display(df, display_map)
    df = attach_targets(df)

    charts = create_charts(df)
    daily_review, planlama_day, operational_day = build_daily_review(df)
    parameter_summaries = build_parameter_summaries(df)
    actions = daily_review
    highlight_actions = build_highlight_actions(actions)

    summary_for_ai = {
        "toplanti_kurali": {
            "isg_kalite_uretim": str(operational_day.date()),
            "planlama": str(planlama_day.date()),
        },
        "gunluk_inceleme": daily_review,
        "parametre_ozetleri": parameter_summaries,
    }

    info_text = (
        f"ISG / Kalite / Uretim icin {operational_day.date()} verileri, "
        f"Planlama icin {planlama_day.date()} verileri yorumlandi."
    )

    return {
        "charts": charts,
        "daily_review": daily_review,
        "actions": actions,
        "highlight_actions": highlight_actions,
        "parameter_summaries": parameter_summaries,
        "summary_for_ai": summary_for_ai,
        "info_text": info_text,
        "operational_day": str(operational_day.date()),
        "planlama_day": str(planlama_day.date()),
    }


def analyze_excel_file(filepath):
    excel_display_map = build_excel_display_map(filepath, sheet_name="Veriler")
    raw_df = pd.read_excel(filepath, sheet_name="Veriler")
    return analyze_raw_dataframe(raw_df, display_map=excel_display_map)
